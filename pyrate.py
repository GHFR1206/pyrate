import os
import pandas as pd
import time
from tabulate import tabulate
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

clear = lambda:os.system('clear')
exit = lambda:os.system('exit')
restart = lambda:os.system('py pyrate.py')

file_path = "data/film.xlsx"
film = pd.read_excel(file_path)
wb = load_workbook(file_path)
ws = wb.active
df = pd.DataFrame(film)
film.index += 1


def movie_list():
    film = pd.read_excel(file_path)
    film.index += 1
    print(tabulate(film, headers='keys', tablefmt='grid'))

def create(name, year, genre):
    data = {
        'id': len(film)+1,
        'name': [name.title()],
        'year': [year],
        'genre': genre,
        'watched': 'No',
        'rate': '-',
        'review': '-',
        }
    data_new = pd.DataFrame(data)
    create = pd.concat([film, data_new], ignore_index=False)
    create.to_excel(file_path, index=False)
    
def edit(id, name, year, genre):
    df.at[id-1, 'name'] = name
    df.at[id-1, 'year'] = year
    df.at[id-1, 'genre'] = genre
    pd.DataFrame(df).to_excel(file_path, index=False)
    
def delete(id):
    df.drop(id-1, inplace=True)
    pd.DataFrame(df).to_excel(file_path, index=False)
    
def search(name):    
    data = df[df["name"].str.contains(name)]
    data.index+=1
    return data

def rating(id, rate, review):
    data = df[df["id"] == id]
    data.index += 1
    rateint = int(rate)
    max = 5
    star = "★" * rateint + "☆" * (max - rateint) + " (" + str(rate) + ")"
    
    data.at[id, 'rate'] = star
    data.at[id, 'watched'] = 'Yes'
    data.at[id, 'review'] = review
    pd.DataFrame(data).to_excel(file_path, index=False)
    
while True:
    clear()
    print("""
██████╗ ██╗   ██╗██████╗  █████╗ ████████╗███████╗
██╔══██╗╚██╗ ██╔╝██╔══██╗██╔══██╗╚══██╔══╝██╔════╝
██████╔╝ ╚████╔╝ ██████╔╝███████║   ██║   █████╗  
██╔═══╝   ╚██╔╝  ██╔══██╗██╔══██║   ██║   ██╔══╝  
██║        ██║   ██║  ██║██║  ██║   ██║   ███████╗
╚═╝        ╚═╝   ╚═╝  ╚═╝╚═╝  ╚═╝   ╚═╝   ╚══════╝
Save your movies here!

what'u want?
[1] Movie list
[0] Exit
        """)
    user = int(input("> "))
    if user == 1:
        clear()
        movie_list()
        print("""
Action:
[1] Add movie
[2] Search movie
[0] Back
              """)
        user = int(input("> "))
        print()
        
        # create
        if user == 1:
            print("insert movie name")
            name = input('> ')
            print("\ninsert movie year")
            year = input('> ')
            print("\ninsert movie genre")
            genre = input('> ')
            create(name, year, genre)
            print("\nCompleted!")
            time.sleep(2)
            continue
                       
        if user == 2:
            print("Type movie name")
            name = input("> ").title()
            data = search(name)
            data = pd.DataFrame(data)
            if search(name).empty:
                print("\nMovie not found!")
                time.sleep(2)
                continue
            
            clear()
            print(tabulate(data, headers='keys', tablefmt='grid'))
            print(f"""
What are you gonna do with that?
[1] Edit
[2] Delete
[3] Rate n Rate it!
[0] Back
                  """)
            user = int(input("> "))
            
            # update
            if user == 1:
                print("Insert new movie name (Leave blank if no change)")
                name = input("> ")
                print("\nInsert new movie year (Leave blank if no change)")
                year = input("> ")
                print("\nInsert new movie genre (Leave blank if no change)")
                genre = input("> ")
                
                if len(name) == 0:
                    name = data.iloc[:]['name'].values[0]
                if len(year) == 0:
                    year = data.iloc[:]['year'].values[0]
                if len(genre) == 0:
                    genre = data.iloc[:]['genre'].values[0]
                            
                edit(data.iloc[:]['id'].values[0], name, year, genre)
                print("\nEdited succesfully!")
                time.sleep(2)
                continue
            
            # delete
            if user == 2:
                print(f"you sure want to delete '{data.at[1, 'name']}'? y/n")
                confirm = input("> ")
                confirm.lower
                if confirm == "y":
                    delete([data.at[1, 'id']])
                    print("Deleted successfully!")
                    time.sleep(2)
                    continue
                else:
                    print("Cancelling..")
                    time.sleep(2)
                    continue
            
            # rating
            if user == 3:
                print("Insert rate number 1-5!")
                rate = float(input("> "))
                if rate > 5:
                    print('Oops.. too much!')
                    time.sleep(1)
                    continue
                print("\nInsert ur review!")
                review = input("> ")
                rating(data.at[1, 'id'], rate, review)
                print("Successfully rated!")
                time.sleep(2)
                continue
            
            if user == 0:
                continue
                        
        if user == 0:
            continue
        
    elif user == 0:
        print("\nBye bye :(")
        exit()
        break
    else:
        print("\nOops, wrong input!")
        time.sleep(2)
        continue
        
